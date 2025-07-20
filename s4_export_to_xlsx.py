import json
import os

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

load_dotenv(override=True)

with open('data_transcript_infer.json', 'r', encoding='utf-8') as file:
    data = json.load(file)

wb = Workbook()
ws = wb.active
ws.title = os.getenv('CHANNEL_NAME')
headers = ['Creator Name', 'Upload Date', 'Title', 'Views', 'Link', 'VPN Mentioned in Title', 'VPN Mentioned in Description', 'AI-Inferred VPN', 'Final Result (Match Between Subtitle and AI Inference)']

ws.append(headers)

for creator, videos in data.items():
    first = True

    for video in videos:
        promoted_vpns = video.get('promoted_vpns', [])
        if promoted_vpns == ['n']: promoted_vpns = []
        final_result = []
        for promoted_vpn in promoted_vpns:
            if promoted_vpn in video.get('transcript', ''):
                final_result.append(promoted_vpn)

        row = [
            creator if first else '',
            video.get('date', ''),
            video.get('title', ''),
            video.get('views', ''),
            video.get('link', ''),
            ', '.join(video.get('vpn_in_title', [])),
            ', '.join(video.get('vpn_in_description', [])),
            ', '.join(promoted_vpns),
            ', '.join(final_result),
        ]
        ws.append(row)
        first = False

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

wb.save('final.xlsx')
